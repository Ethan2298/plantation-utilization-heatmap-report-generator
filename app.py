"""
Plantation Schedule Optimizer
Upload your Zenoti exports → get heatmaps + staffing recommendations instantly.
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import matplotlib.gridspec as gridspec
from bs4 import BeautifulSoup
from datetime import datetime, timedelta, time, date
from collections import defaultdict, Counter
from io import BytesIO

st.set_page_config(page_title="Plantation Schedule Optimizer", layout="wide")

# ── Helpers ──

SLOT_MINUTES = 30
DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

def build_slots(slot_start, slot_end):
    slots = []
    t = datetime(2000, 1, 1, slot_start.hour, slot_start.minute)
    end = datetime(2000, 1, 1, slot_end.hour, slot_end.minute)
    while t < end:
        slots.append(t.time())
        t += timedelta(minutes=SLOT_MINUTES)
    return slots

def time_to_minutes(t):
    return t.hour * 60 + t.minute

def slot_overlaps(slot_start, interval_start, interval_end):
    s1 = time_to_minutes(slot_start)
    s2 = s1 + SLOT_MINUTES
    i1 = time_to_minutes(interval_start)
    i2 = time_to_minutes(interval_end)
    return s1 < i2 and i1 < s2

def parse_time_str(s):
    s = s.strip().upper()
    for fmt in ("%I:%M%p", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Cannot parse time: {s}")

def parse_schedule(sched):
    parts = sched.split("-")
    if len(parts) != 2:
        return None, None
    return parse_time_str(parts[0]), parse_time_str(parts[1])

def slot_label(s):
    h = s.hour; m = s.minute
    ampm = "AM" if h < 12 else "PM"
    h12 = h if h <= 12 else h - 12
    if h12 == 0: h12 = 12
    return f"{h12}:{m:02d} {ampm}"


# ── Parse functions ──

def parse_attendance(file, date_start, date_end):
    """Parse attendance XLSX → {(date, employee): (shift_start, shift_end)}"""
    import openpyxl
    wb = openpyxl.load_workbook(file, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    shifts = {}
    for row in sheet.iter_rows(min_row=5, values_only=True):
        date_val, emp_name, job, status, center, schedule = row[0], row[1], row[2], row[3], row[4], row[5]
        if status != "Working" or not schedule:
            continue
        if isinstance(date_val, datetime):
            dt = date_val.date()
        elif isinstance(date_val, str):
            try:
                dt = datetime.strptime(date_val, "%m/%d/%Y").date()
            except ValueError:
                continue
        else:
            continue
        if dt < date_start or dt > date_end:
            continue
        shift_start, shift_end = parse_schedule(schedule)
        if shift_start and shift_end:
            shifts[(dt, emp_name)] = (shift_start, shift_end)
    return shifts


def parse_blockouts(file, date_start, date_end, exclude_types):
    """Parse block-out HTML/XLS → {(date, employee): [(start, end), ...]}"""
    content = file.read().decode("utf-8", errors="ignore")
    soup = BeautifulSoup(content, "html.parser")
    blockouts = defaultdict(list)
    excluded = 0
    included = 0
    for tr in soup.find_all("tr"):
        cells = tr.find_all("td")
        if len(cells) < 8:
            continue
        emp_raw = cells[1].get_text(strip=True)
        date_str = cells[2].get_text(strip=True)
        block_type = cells[5].get_text(strip=True)
        start_str = cells[6].get_text(strip=True)
        end_str = cells[7].get_text(strip=True)
        if not date_str or not start_str or not end_str:
            continue
        if block_type in exclude_types:
            excluded += 1
            continue
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            continue
        if dt < date_start or dt > date_end:
            continue
        try:
            bo_start = parse_time_str(start_str)
            bo_end = parse_time_str(end_str)
        except ValueError:
            continue
        emp_name = emp_raw.split("(")[0].strip()
        blockouts[(dt, emp_name)].append((bo_start, bo_end))
        included += 1
    return blockouts, included, excluded


def parse_blockouts_all(file_bytes, date_start, date_end):
    """Parse ALL block-outs (no exclusions) for supply analysis."""
    soup = BeautifulSoup(file_bytes.decode("utf-8", errors="ignore"), "html.parser")
    blockouts = defaultdict(list)
    for tr in soup.find_all("tr"):
        cells = tr.find_all("td")
        if len(cells) < 8:
            continue
        emp_raw = cells[1].get_text(strip=True)
        date_str = cells[2].get_text(strip=True)
        start_str = cells[6].get_text(strip=True)
        end_str = cells[7].get_text(strip=True)
        if not date_str or not start_str or not end_str:
            continue
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            continue
        if dt < date_start or dt > date_end:
            continue
        try:
            bo_start = parse_time_str(start_str)
            bo_end = parse_time_str(end_str)
        except ValueError:
            continue
        emp_name = emp_raw.split("(")[0].strip()
        blockouts[(dt, emp_name)].append((bo_start, bo_end))
    return blockouts


def get_block_types(file_bytes):
    """Get unique block-out types from the file."""
    soup = BeautifulSoup(file_bytes.decode("utf-8", errors="ignore"), "html.parser")
    types = set()
    for tr in soup.find_all("tr"):
        cells = tr.find_all("td")
        if len(cells) >= 6:
            t = cells[5].get_text(strip=True)
            if t:
                types.add(t)
    return sorted(types)


def parse_appointments(file, date_start, date_end):
    """Parse appointments XLSX → {(date, therapist): [(start, end), ...]}"""
    import openpyxl
    wb = openpyxl.load_workbook(file, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    appointments = defaultdict(list)
    count = 0
    for row in sheet.iter_rows(min_row=5, values_only=True):
        appt_date, start_val, end_val = row[0], row[6], row[7]
        therapist, service_cat = row[11], row[16]
        if service_cat != "Massages" or not therapist or not start_val or not end_val:
            continue
        if isinstance(appt_date, datetime):
            dt = appt_date.date()
        elif isinstance(appt_date, str):
            try:
                dt = datetime.strptime(appt_date, "%m/%d/%Y").date()
            except ValueError:
                continue
        else:
            continue
        if dt < date_start or dt > date_end:
            continue
        if isinstance(start_val, datetime):
            a_start = start_val.time()
        elif isinstance(start_val, str):
            try:
                a_start = datetime.strptime(start_val, "%m/%d/%Y %I:%M %p").time()
            except ValueError:
                continue
        else:
            continue
        if isinstance(end_val, datetime):
            a_end = end_val.time()
        elif isinstance(end_val, str):
            try:
                a_end = datetime.strptime(end_val, "%m/%d/%Y %I:%M %p").time()
            except ValueError:
                continue
        else:
            continue
        if a_start == a_end:
            continue
        appointments[(dt, therapist)].append((a_start, a_end))
        count += 1
    return appointments, count


# ── Calculation engine ──

def calculate_idle_heatmap(shifts, blockouts, appointments, date_start, date_end, slots):
    """Calculate idle therapist-slots heatmap by day of week."""
    heatmap = np.zeros((7, len(slots)), dtype=int)
    cur = date_start
    while cur <= date_end:
        dow = cur.weekday()
        emps = {emp: s for (d, emp), s in shifts.items() if d == cur}
        for si, slot_time in enumerate(slots):
            idle = 0
            for emp, (ss, se) in emps.items():
                if not slot_overlaps(slot_time, ss, se):
                    continue
                blocked = any(slot_overlaps(slot_time, bs, be) for bs, be in blockouts.get((cur, emp), []))
                if blocked:
                    continue
                in_appt = any(slot_overlaps(slot_time, a1, a2) for a1, a2 in appointments.get((cur, emp), []))
                if in_appt:
                    continue
                idle += 1
            heatmap[dow][si] += idle
        cur += timedelta(days=1)
    return heatmap


def calculate_supply_demand(shifts, blockouts_all, appointments, date_start, date_end, slots):
    """Calculate supply/demand/utilization matrices (weekly averages)."""
    day_counts = Counter()
    cur = date_start
    while cur <= date_end:
        day_counts[cur.weekday()] += 1
        cur += timedelta(days=1)

    supply_raw = np.zeros((7, len(slots)), dtype=float)
    supply_avail = np.zeros((7, len(slots)), dtype=float)
    demand = np.zeros((7, len(slots)), dtype=float)

    cur = date_start
    while cur <= date_end:
        dow = cur.weekday()
        emps = {emp: s for (d, emp), s in shifts.items() if d == cur}
        for si, slot_time in enumerate(slots):
            for emp, (ss, se) in emps.items():
                if not slot_overlaps(slot_time, ss, se):
                    continue
                supply_raw[dow][si] += 1
                blocked = any(slot_overlaps(slot_time, bs, be) for bs, be in blockouts_all.get((cur, emp), []))
                if blocked:
                    continue
                supply_avail[dow][si] += 1
                in_appt = any(slot_overlaps(slot_time, a1, a2) for a1, a2 in appointments.get((cur, emp), []))
                if in_appt:
                    demand[dow][si] += 1
        cur += timedelta(days=1)

    for dow in range(7):
        n = day_counts[dow]
        if n > 0:
            supply_raw[dow] /= n
            supply_avail[dow] /= n
            demand[dow] /= n

    with np.errstate(divide='ignore', invalid='ignore'):
        utilization = np.where(supply_avail > 0, demand / supply_avail * 100, 0)
    idle_avg = supply_avail - demand

    return supply_raw, supply_avail, demand, utilization, idle_avg, day_counts


# ── Chart builders ──

def build_idle_heatmap_fig(heatmap, slots, title_suffix="", subtitle_extra=""):
    labels = [slot_label(s) for s in slots]
    row_totals = heatmap.sum(axis=1)
    total = heatmap.sum()

    fig, ax = plt.subplots(figsize=(22, 5.5))
    cmap = mcolors.LinearSegmentedColormap.from_list("idle_red", ["#fef0ef", "#67000d"])
    im = ax.imshow(heatmap, cmap=cmap, aspect="auto", vmin=0, vmax=max(20, heatmap.max()))

    ax.set_yticks(range(7))
    ax.set_yticklabels(DAY_ORDER, fontsize=12, fontweight="bold")
    ax.set_xticks(range(len(slots)))
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=9)

    for i in range(7):
        for j in range(len(slots)):
            v = heatmap[i][j]
            if v > 0:
                ax.text(j, i, str(v), ha="center", va="center", fontsize=9,
                        fontweight="bold", color="white" if v > 12 else "black")

    for i in range(7):
        ax.text(len(slots) + 0.3, i, str(int(row_totals[i])),
                ha="left", va="center", fontsize=13, fontweight="bold")
    ax.text(len(slots) + 0.3, -0.8, "Total", ha="left", va="center", fontsize=11, fontweight="bold")

    fig.suptitle(f"Plantation Therapist Idle Time{title_suffix}", fontsize=16, fontweight="bold", y=0.98)
    ax.set_title(f"{total:,} idle therapist-slots  |  {subtitle_extra}", fontsize=10, color="gray", pad=12)

    plt.colorbar(im, ax=ax, shrink=0.8, pad=0.08).set_label("Idle Therapist-Slots", fontsize=10)
    plt.tight_layout(rect=[0, 0, 0.93, 0.92])
    return fig


def build_optimization_fig(supply_avail, demand, utilization, idle_avg, slots):
    labels = [slot_label(s) for s in slots]

    fig = plt.figure(figsize=(24, 20))
    gs = gridspec.GridSpec(4, 1, height_ratios=[1, 1, 1, 1], hspace=0.35)

    cmaps = [
        mcolors.LinearSegmentedColormap.from_list("s", ["#f0f4ff", "#1a3a6b"]),
        mcolors.LinearSegmentedColormap.from_list("d", ["#f0fff0", "#1a6b2b"]),
        mcolors.LinearSegmentedColormap.from_list("u", ["#f0fff0", "#ffff00", "#ff6600", "#cc0000"]),
        mcolors.LinearSegmentedColormap.from_list("i", ["#fef0ef", "#67000d"]),
    ]
    panels = [
        (supply_avail, "Avg Available Therapists (After Block-Outs)", cmaps[0], 0, 10, ".1f"),
        (demand, "Avg Massage Bookings (Therapists Occupied)", cmaps[1], 0, 8, ".1f"),
        (utilization, "Utilization Rate (%)", cmaps[2], 0, 100, ".0f"),
        (idle_avg, "Avg Idle Therapists (Opportunity to Optimize)", cmaps[3], 0, 8, ".1f"),
    ]

    for idx, (data, title, cmap, vmin, vmax, fmt) in enumerate(panels):
        ax = fig.add_subplot(gs[idx])
        im = ax.imshow(data, cmap=cmap, aspect="auto", vmin=vmin, vmax=vmax)
        ax.set_yticks(range(7))
        ax.set_yticklabels(DAY_ORDER, fontsize=11, fontweight="bold")
        ax.set_xticks(range(len(slots)))
        ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
        ax.set_title(title, fontsize=13, fontweight="bold", pad=8)

        for i in range(7):
            for j in range(len(slots)):
                v = data[i][j]
                if v > 0.05:
                    brightness = (v - vmin) / (vmax - vmin) if vmax > vmin else 0
                    ax.text(j, i, f"{v:{fmt}}", ha="center", va="center", fontsize=7,
                            fontweight="bold", color="white" if brightness > 0.55 else "black")

        row_avgs = data.mean(axis=1)
        for i in range(7):
            ax.text(len(slots) + 0.3, i, f"{row_avgs[i]:{fmt}}",
                    ha="left", va="center", fontsize=10, fontweight="bold")
        ax.text(len(slots) + 0.3, -0.8, "Avg", ha="left", va="center", fontsize=9, fontweight="bold", color="gray")
        plt.colorbar(im, ax=ax, shrink=0.7, pad=0.06)

    fig.suptitle("Schedule Optimization - Supply vs. Demand Analysis",
                 fontsize=16, fontweight="bold", y=0.98)
    return fig


def build_recommendations(supply_avail, demand, utilization, idle_avg, slots):
    """Build add/cut recommendations as DataFrames."""
    labels = [slot_label(s) for s in slots]

    cut_rows = []
    add_rows = []
    for dow in range(7):
        for si in range(len(slots)):
            avail = supply_avail[dow][si]
            dem = demand[dow][si]
            util = utilization[dow][si]
            idle_v = idle_avg[dow][si]
            if avail < 0.5:
                continue
            row = {
                "Day": DAY_ORDER[dow],
                "Time": labels[si],
                "Avg Available": round(avail, 1),
                "Avg Booked": round(dem, 1),
                "Avg Idle": round(idle_v, 1),
                "Utilization": f"{util:.0f}%"
            }
            if util >= 85 and dem >= 1:
                add_rows.append({**row, "Priority": "High" if util >= 95 else "Medium"})
            if idle_v >= 1.2 and util < 75:
                cut_rows.append(row)

    add_df = pd.DataFrame(add_rows).sort_values("Utilization", ascending=False) if add_rows else pd.DataFrame()
    cut_df = pd.DataFrame(cut_rows).sort_values("Avg Idle", ascending=False) if cut_rows else pd.DataFrame()
    return add_df, cut_df


def fig_to_pdf_bytes(fig):
    buf = BytesIO()
    fig.savefig(buf, format="pdf", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════

st.title("Plantation Schedule Optimizer")
st.markdown("Upload your 3 Zenoti exports and get instant heatmaps + staffing recommendations.")

st.divider()

col1, col2, col3 = st.columns(3)
with col1:
    att_file = st.file_uploader("Attendance (.xlsx)", type=["xlsx"])
with col2:
    bo_file = st.file_uploader("Block Out Time Details (.xls)", type=["xls"])
with col3:
    appt_file = st.file_uploader("Appointments (.xlsx)", type=["xlsx"])

if att_file and bo_file and appt_file:
    # Detect date range from attendance
    st.divider()

    # Read block-out file bytes once (since we need it multiple times)
    bo_bytes = bo_file.read()
    bo_file.seek(0)

    # Get block types for the exclusion selector
    block_types = get_block_types(bo_bytes)

    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("Date Range")
        date_start = st.date_input("Start date", value=date(2025, 12, 1))
        date_end = st.date_input("End date", value=date(2026, 2, 9))
    with col_b:
        st.subheader("Block-Out Types to EXCLUDE from Idle Calc")
        st.caption("These block types will NOT count as valid block-outs. Time covered by them will show as idle.")
        exclude_types = st.multiselect(
            "Select types to exclude",
            options=block_types,
            default=[t for t in block_types if "Leaving Early" in t or "Shift Adjustment" in t]
        )

    if st.button("Generate Analysis", type="primary", use_container_width=True):
        with st.spinner("Parsing data..."):
            shifts = parse_attendance(att_file, date_start, date_end)
            blockouts_filtered, inc, exc = parse_blockouts(
                BytesIO(bo_bytes), date_start, date_end, set(exclude_types)
            )
            blockouts_all = parse_blockouts_all(bo_bytes, date_start, date_end)
            appointments, appt_count = parse_appointments(appt_file, date_start, date_end)

        st.success(f"Parsed {len(shifts)} shifts, {inc} block-outs ({exc} excluded), {appt_count} massage appointments")

        slots = build_slots(time(9, 0), time(21, 0))

        # ── Tab 1: Idle Heatmap (with exclusions) ──
        # ── Tab 2: Idle Heatmap (all block-outs) ──
        # ── Tab 3: Supply vs Demand ──
        # ── Tab 4: Recommendations ──

        tab1, tab2, tab3, tab4 = st.tabs([
            "Idle Heatmap (Adjusted)",
            "Idle Heatmap (Original)",
            "Supply vs. Demand",
            "Recommendations"
        ])

        with tab1:
            with st.spinner("Calculating adjusted idle heatmap..."):
                hm_adj = calculate_idle_heatmap(shifts, blockouts_filtered, appointments, date_start, date_end, slots)
            excl_label = " & ".join(t.replace(" 2026", "") for t in exclude_types) if exclude_types else "None"
            fig1 = build_idle_heatmap_fig(
                hm_adj, slots,
                title_suffix=f" - Minus {excl_label} Blocks",
                subtitle_extra=f"Excluding: {excl_label}"
            )
            st.pyplot(fig1)
            st.download_button("Download PDF", fig_to_pdf_bytes(fig1),
                               file_name="Idle_Heatmap_Adjusted.pdf", mime="application/pdf")

        with tab2:
            with st.spinner("Calculating original idle heatmap..."):
                hm_orig = calculate_idle_heatmap(shifts, blockouts_all, appointments, date_start, date_end, slots)
            fig2 = build_idle_heatmap_fig(
                hm_orig, slots,
                title_suffix="",
                subtitle_extra="All block-outs included"
            )
            st.pyplot(fig2)
            st.download_button("Download PDF", fig_to_pdf_bytes(fig2),
                               file_name="Idle_Heatmap_Original.pdf", mime="application/pdf")

        with tab3:
            with st.spinner("Calculating supply vs. demand..."):
                sr, sa, dem, util, idle_a, dc = calculate_supply_demand(
                    shifts, blockouts_all, appointments, date_start, date_end, slots
                )
            fig3 = build_optimization_fig(sa, dem, util, idle_a, slots)
            st.pyplot(fig3)
            st.download_button("Download PDF", fig_to_pdf_bytes(fig3),
                               file_name="Supply_vs_Demand.pdf", mime="application/pdf")

        with tab4:
            with st.spinner("Building recommendations..."):
                sr, sa, dem, util, idle_a, dc = calculate_supply_demand(
                    shifts, blockouts_all, appointments, date_start, date_end, slots
                )
                add_df, cut_df = build_recommendations(sa, dem, util, idle_a, slots)

            st.subheader("Where to ADD Hours (85%+ utilization)")
            st.caption("These slots are near or at capacity. Consider adding therapist hours.")
            if not add_df.empty:
                st.dataframe(add_df, use_container_width=True, hide_index=True)
            else:
                st.info("No slots at 85%+ utilization.")

            st.divider()

            st.subheader("Where to CUT Hours (1.2+ avg idle therapists, <75% utilization)")
            st.caption("These slots consistently have idle therapists. Consider reducing hours.")
            if not cut_df.empty:
                st.dataframe(cut_df, use_container_width=True, hide_index=True)
            else:
                st.info("No significantly overstaffed slots found.")

else:
    st.info("Upload all 3 files above to get started.")
